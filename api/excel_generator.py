#!/usr/bin/env python3
"""
Excel Generator for Order AWB Lists

Generates Excel files with order AWBs, grouping products and separating complex orders.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple
from collections import defaultdict
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import get_config
from logging_config import get_logger


class ExcelGenerator:
    """Generates Excel files for order AWB lists."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.config = get_config()
        self.logger = get_logger('excel')
        self.output_dir = self.config.output_dir
        
        # Load Excel formatting configuration
        self.format_config = self._load_format_config()
        
        self.logger.info("Excel generator initialized")
    
    def _load_format_config(self) -> Dict:
        """
        Load Excel formatting configuration from JSON file.
        
        Returns:
            Dictionary with formatting configuration
        """
        config_file = self.config.data_dir / 'excel_format_config.json'
        
        # Default configuration
        default_config = {
            "columns": {
                "Nr.": {"width": 8, "font_size": 11, "bold": False},
                "Nume Client": {"width": 25, "font_size": 11, "bold": False},
                "Cantitate": {"width": 10, "font_size": 11, "bold": False},
                "Nume Produs": {"width": 35, "font_size": 11, "bold": False},
                "Culoare": {"width": 15, "font_size": 11, "bold": False},
                "Cod Produs": {"width": 12, "font_size": 11, "bold": False},
                "AWB": {"width": 18, "font_size": 11, "bold": False}
            },
            "header": {
                "font_size": 12,
                "bold": True,
                "background_color": "4472C4",
                "font_color": "FFFFFF"
            },
            "separator": {
                "font_size": 14,
                "bold": True,
                "font_color": "FF0000",
                "background_color": "FFE699"
            },
            "formatting": {
                "add_empty_row_between_products": True,
                "add_borders": True
            }
        }
        
        try:
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.logger.info(f"Loaded Excel format config from {config_file}")
                return config
            else:
                self.logger.warning(f"Excel format config not found at {config_file}, using defaults")
                # Create default config file
                with open(config_file, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, indent=2, ensure_ascii=False)
                self.logger.info(f"Created default Excel format config at {config_file}")
                return default_config
        except Exception as e:
            self.logger.error(f"Error loading Excel format config: {e}, using defaults")
            return default_config
    
    def _identify_complex_clients(self, orders: List[Dict]) -> set:
        """
        Identify clients that should be in the "Comenzi complexe" section.
        
        Args:
            orders: List of order dictionaries with 'customer_name' and 'quantity'
            
        Returns:
            Set of client names that are "complex"
        """
        client_order_count = defaultdict(int)
        client_has_multiple_qty = set()
        
        # Count orders per client and check quantities
        for order in orders:
            client_name = order.get('customer_name', '').strip()
            quantity = order.get('quantity', 1)
            
            client_order_count[client_name] += 1
            
            if quantity > 1:
                client_has_multiple_qty.add(client_name)
        
        # Complex clients: appear more than once OR have any order with qty > 1
        complex_clients = set()
        
        for client_name, count in client_order_count.items():
            if count > 1 or client_name in client_has_multiple_qty:
                complex_clients.add(client_name)
        
        self.logger.info(f"Identified {len(complex_clients)} complex clients out of {len(client_order_count)} total")
        
        return complex_clients
    
    def _group_and_sort_orders(self, orders: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
        """
        Group identical products and separate simple vs complex orders.
        
        Args:
            orders: List of order dictionaries
            
        Returns:
            Tuple of (simple_orders, complex_orders)
        """
        # Identify complex clients
        complex_clients = self._identify_complex_clients(orders)
        
        simple_orders = []
        complex_orders = []
        
        # Separate orders
        for order in orders:
            client_name = order.get('customer_name', '').strip()
            
            if client_name in complex_clients:
                complex_orders.append(order)
            else:
                simple_orders.append(order)
        
        # Sort simple orders by product name for grouping
        simple_orders.sort(key=lambda x: (x.get('product_name', ''), x.get('color', '')))
        
        # Sort complex orders by CLIENT NAME first, then product name
        # This keeps all orders from the same client together
        complex_orders.sort(key=lambda x: (x.get('customer_name', ''), x.get('product_name', ''), x.get('color', '')))
        
        self.logger.info(f"Grouped orders: {len(simple_orders)} simple, {len(complex_orders)} complex")
        
        return simple_orders, complex_orders
    
    def _create_excel_workbook(self, simple_orders: List[Dict], complex_orders: List[Dict],
                               filename: str = None) -> Path:
        """
        Create Excel workbook with orders.
        
        Args:
            simple_orders: List of simple orders
            complex_orders: List of complex orders
            filename: Output filename (auto-generated if None)
            
        Returns:
            Path to created Excel file
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders AWB List"
        
        # Get formatting config
        header_config = self.format_config.get('header', {})
        separator_config = self.format_config.get('separator', {})
        columns_config = self.format_config.get('columns', {})
        formatting_config = self.format_config.get('formatting', {})
        
        # Define styles from config
        header_font = Font(
            bold=header_config.get('bold', True),
            size=header_config.get('font_size', 12),
            color=header_config.get('font_color', 'FFFFFF')
        )
        header_fill = PatternFill(
            start_color=header_config.get('background_color', '4472C4'),
            end_color=header_config.get('background_color', '4472C4'),
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        separator_font = Font(
            bold=separator_config.get('bold', True),
            size=separator_config.get('font_size', 14),
            color=separator_config.get('font_color', 'FF0000')
        )
        separator_fill = PatternFill(
            start_color=separator_config.get('background_color', 'FFE699'),
            end_color=separator_config.get('background_color', 'FFE699'),
            fill_type="solid"
        )
        separator_alignment = Alignment(horizontal="left", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ) if formatting_config.get('add_borders', True) else None
        
        # Write header and set column widths
        headers = ['Nr.', 'Nume Client', 'Cantitate', 'Nume Produs', 'Culoare', 'Cod Produs']
        for col_idx, header in enumerate(headers, start=1):
            # Get column config
            col_config = columns_config.get(header, {})
            col_letter = get_column_letter(col_idx)
            
            # Set column width
            ws.column_dimensions[col_letter].width = col_config.get('width', 15)
            
            # Write header cell
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            if border_style:
                cell.border = border_style
        
        # Set print titles - repeat header row on every printed page
        ws.print_title_rows = '1:1'  # Repeat row 1 on every page
        
        current_row = 2
        order_number = 1
        
        # Write simple orders with product grouping
        self.logger.info(f"Writing {len(simple_orders)} simple orders...")
        previous_product = None
        add_separator = formatting_config.get('add_empty_row_between_products', True)
        
        for order in simple_orders:
            # Check if product changed - add empty row between different products
            current_product = (order.get('product_name', ''), order.get('color', ''))
            if add_separator and previous_product is not None and previous_product != current_product:
                # Add empty row between different products
                current_row += 1
            previous_product = current_product
            
            # Write order data
            values = [
                order_number,
                order.get('customer_name', ''),
                order.get('quantity', 1),
                order.get('product_name', ''),
                order.get('color', ''),
                order.get('razz_code', '')
            ]
            
            for col_idx, (value, header) in enumerate(zip(values, headers), start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                
                # Apply column-specific formatting
                col_config = columns_config.get(header, {})
                cell.font = Font(
                    size=col_config.get('font_size', 11),
                    bold=col_config.get('bold', False)
                )
                
                # Apply center alignment for Quantity, Color, and Product Code columns
                if header in ['Cantitate', 'Culoare', 'Cod Produs']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply borders
                if border_style:
                    cell.border = border_style
            
            order_number += 1
            current_row += 1
        
        # Add separator if there are complex orders
        if complex_orders:
            # Empty row
            current_row += 1
            
            # Separator row
            separator_cell = ws.cell(row=current_row, column=1, value="Comenzi complexe")
            separator_cell.font = separator_font
            separator_cell.fill = separator_fill
            separator_cell.alignment = separator_alignment
            
            # Merge cells for separator (now includes AWB column)
            ws.merge_cells(f'A{current_row}:F{current_row}')
            
            # Apply border to merged cell
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border_style
            
            current_row += 1
            
            # Empty row after separator
            current_row += 1
            
            # Add AWB column header for complex orders section (after Cod Produs)
            awb_header_text = 'AWB'
            awb_config = columns_config.get(awb_header_text, {})
            awb_header = ws.cell(row=current_row, column=7, value=awb_header_text)
            awb_header.font = header_font
            awb_header.fill = header_fill
            awb_header.alignment = header_alignment
            if border_style:
                awb_header.border = border_style
            
            # Set AWB column width from config
            ws.column_dimensions['G'].width = awb_config.get('width', 18)
            
            current_row += 1
            
            # Track previous product for separators in complex section too
            previous_complex_product = None
            previous_customer = None
            
            # Identify customers with multiple different AWBs
            customer_awbs = {}
            for order in complex_orders:
                customer = order.get('customer_name', '')
                awb = order.get('cargo_tracking_number', '')
                if customer not in customer_awbs:
                    customer_awbs[customer] = set()
                customer_awbs[customer].add(awb)
            
            # Customers with different AWBs (multiple separate orders)
            multi_awb_customers = {customer for customer, awbs in customer_awbs.items() if len(awbs) > 1}
            
            # Identify customers with multiple different products
            customer_products = {}
            for order in complex_orders:
                customer = order.get('customer_name', '')
                product = (order.get('product_name', ''), order.get('color', ''))
                if customer not in customer_products:
                    customer_products[customer] = set()
                customer_products[customer].add(product)
            
            # Customers with multiple different products (need rectangle border)
            multi_product_customers = {customer for customer, products in customer_products.items() if len(products) > 1}
            
            # Debug: log multi-product customers
            self.logger.info(f"Multi-product customers detected: {multi_product_customers}")
            for customer, products in customer_products.items():
                if len(products) > 1:
                    self.logger.info(f"  {customer}: {len(products)} products - {products}")
            
            # Write complex orders
            self.logger.info(f"Writing {len(complex_orders)} complex orders...")
            complex_headers = headers + [awb_header_text]
            
            # Track customer row ranges for rectangle borders
            customer_row_ranges = {}
            current_customer_start_row = None
            
            for order in complex_orders:
                customer = order.get('customer_name', '')
                is_multi_awb = customer in multi_awb_customers
                is_multi_product = customer in multi_product_customers
                
                # Add empty row when customer changes (to separate different customers)
                if add_separator and previous_customer is not None and previous_customer != customer:
                    # End previous customer's range if they had multiple products
                    if previous_customer in multi_product_customers and current_customer_start_row is not None:
                        customer_row_ranges[previous_customer] = (current_customer_start_row, current_row - 1)
                        self.logger.info(f"Recorded range for {previous_customer}: rows {current_customer_start_row}-{current_row - 1}")
                    
                    current_row += 1
                    # Start new customer range
                    if customer in multi_product_customers:
                        current_customer_start_row = current_row
                    else:
                        current_customer_start_row = None
                elif previous_customer != customer:
                    # First customer or customer change without separator
                    if customer in multi_product_customers:
                        current_customer_start_row = current_row
                    else:
                        current_customer_start_row = None
                
                previous_customer = customer
                
                # Write order data
                values = [
                    order_number,
                    customer,
                    order.get('quantity', 1),
                    order.get('product_name', ''),
                    order.get('color', ''),
                    order.get('razz_code', ''),
                    order.get('cargo_tracking_number', '')
                ]
                
                for col_idx, (value, header) in enumerate(zip(values, complex_headers), start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    
                    # Get column config
                    col_config = columns_config.get(header, {})
                    
                    # Special formatting for quantity > 1 (bold and red)
                    if header == 'Cantitate' and value > 1:
                        cell.font = Font(
                            color="FF0000",
                            size=col_config.get('font_size', 11),
                            bold=True
                        )
                    # Apply RED font ONLY to AWB column if customer has multiple different AWBs
                    elif header == 'AWB' and is_multi_awb:
                        cell.font = Font(
                            color="FF0000",
                            size=col_config.get('font_size', 11),
                            bold=False
                        )
                    else:
                        # Apply column-specific formatting
                        cell.font = Font(
                            size=col_config.get('font_size', 11),
                            bold=col_config.get('bold', False)
                        )
                    
                    # Apply center alignment for Quantity, Color, and Product Code columns
                    if header in ['Cantitate', 'Culoare', 'Cod Produs', 'AWB']:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Apply borders
                    if border_style:
                        cell.border = border_style
                
                order_number += 1
                current_row += 1
            
            # Close the last customer's range if they had multiple products
            if previous_customer in multi_product_customers and current_customer_start_row is not None:
                customer_row_ranges[previous_customer] = (current_customer_start_row, current_row - 1)
                self.logger.info(f"Recorded final range for {previous_customer}: rows {current_customer_start_row}-{current_row - 1}")
            
            # Apply rectangle borders for customers with multiple products
            # Use medium thick borders for better visibility (double wasn't prominent enough)
            extra_thick_border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
            
            for customer, (start_row, end_row) in customer_row_ranges.items():
                self.logger.info(f"Applying rectangle border for customer {customer}: rows {start_row}-{end_row}")
                # Apply extra thick border to all cells in the customer's row range
                for row in range(start_row, end_row + 1):
                    for col in range(1, 8):  # Columns A to G (including AWB)
                        cell = ws.cell(row=row, column=col)
                        
                        # Determine which borders to apply based on position in rectangle (using thick + blue for visibility)
                        top_border = Side(style='thick', color='0000FF') if row == start_row else None
                        bottom_border = Side(style='thick', color='0000FF') if row == end_row else None
                        left_border = Side(style='thick', color='0000FF') if col == 1 else None
                        right_border = Side(style='thick', color='0000FF') if col == 7 else None
                        
                        # Preserve existing thin borders for internal cells, add thick for rectangle outline
                        existing_border = cell.border
                        new_border = Border(
                            left=left_border if left_border else existing_border.left,
                            right=right_border if right_border else existing_border.right,
                            top=top_border if top_border else existing_border.top,
                            bottom=bottom_border if bottom_border else existing_border.bottom
                        )
                        cell.border = new_border
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = f"{timestamp}_orders.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Save file
        output_path = self.output_dir / filename
        wb.save(output_path)
        
        self.logger.info(f"Excel file created: {output_path}")
        return output_path
    
    def generate_from_orders(self, orders: List[Dict], filename: str = None) -> Path:
        """
        Generate Excel file from list of orders.
        
        Args:
            orders: List of order dictionaries with keys:
                   - customer_name: Client full name
                   - quantity: Item quantity
                   - product_name: Simplified product name
                   - color: Extracted color
            filename: Optional custom filename
            
        Returns:
            Path to generated Excel file
        """
        if not orders:
            raise ValueError("No orders provided to generate Excel file")
        
        self.logger.info(f"Generating Excel file from {len(orders)} orders")
        
        # Group and sort orders
        simple_orders, complex_orders = self._group_and_sort_orders(orders)
        
        # Create Excel file
        output_path = self._create_excel_workbook(simple_orders, complex_orders, filename)
        
        self.logger.info(f"Successfully generated Excel file: {output_path}")
        self.logger.info(f"  - Simple orders: {len(simple_orders)}")
        self.logger.info(f"  - Complex orders: {len(complex_orders)}")
        self.logger.info(f"  - Total orders: {len(orders)}")
        
        # Save Excel order (simple + complex in display order)
        all_orders = simple_orders + complex_orders
        excel_order_path = output_path.parent / output_path.name.replace('.xlsx', '_excel_order.json')
        with open(excel_order_path, 'w', encoding='utf-8') as f:
            json.dump(all_orders, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"  - Excel order saved for AWB reordering")
        
        return output_path


def test_excel_generator():
    """Test the Excel generator with sample data."""
    print("Testing Excel Generator")
    print("=" * 60)
    
    # Sample orders
    sample_orders = [
        {'customer_name': 'John Doe', 'quantity': 1, 'product_name': 'Blender SilverCrest', 'color': 'Argintiu'},
        {'customer_name': 'John Doe', 'quantity': 2, 'product_name': 'Raft metalic', 'color': 'Negru'},
        {'customer_name': 'Jane Smith', 'quantity': 1, 'product_name': 'Blender SilverCrest', 'color': 'Argintiu'},
        {'customer_name': 'Mike Brown', 'quantity': 1, 'product_name': 'Cos gunoi Smart', 'color': 'Alb'},
        {'customer_name': 'Anna White', 'quantity': 3, 'product_name': 'Cos gunoi Smart', 'color': 'Roz'},
        {'customer_name': 'Bob Green', 'quantity': 1, 'product_name': 'Raft metalic', 'color': 'Negru'},
    ]
    
    generator = ExcelGenerator()
    
    try:
        output_file = generator.generate_from_orders(sample_orders, "test_orders.xlsx")
        print(f"\n✅ Excel file generated successfully!")
        print(f"   Location: {output_file}")
    except Exception as e:
        print(f"\n❌ Error generating Excel: {e}")
    
    print("=" * 60)


if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO)
    
    test_excel_generator()

